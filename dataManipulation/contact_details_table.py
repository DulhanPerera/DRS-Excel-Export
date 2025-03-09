# contact_details_table.py
import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import logging

def create_contact_details_table(ws, case_data, start_row, start_col):
    """
    Create the Contact Details table in the worksheet.
    """
    try:
        logging.info("Creating Contact Details table...")
        contacts_headers = ["Mobile", "Email", "Home Phone", "Address"]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
        ws.cell(row=start_row, column=start_col, value="Contact Info").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(contacts_headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert contact data
        contacts = case_data.get("contact", [])
        for contact_idx, contact in enumerate(contacts, start=1):
            ws.cell(row=start_row + 1 + contact_idx, column=start_col, value=contact.get("mob")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 1, value=contact.get("email")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 2, value=contact.get("lan")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 3, value=contact.get("address")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(contacts_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(contacts) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logging.info("Contact Details table created successfully.")
        return start_row + len(contacts) + 3
    except Exception as e:
        logging.error(f"Failed to create Contact Details table: {e}")
        sys.exit(1)