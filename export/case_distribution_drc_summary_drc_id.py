import logging
from datetime import datetime
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
import os
from pymongo import MongoClient

logger = logging.getLogger('excel_data_writer')

DRC_SUMMARY_HEADERS = [
    "created_dtm", "drc_id", "drc", "case_count", "tot_arrease", "proceed_on"
]

def excel_drc_summary_detail(drc, case_distribution_batch_id):
    """Fetch and export DRC summary details with a fixed Task_Id of 20 based on validated parameters"""
    
    try:
        client = MongoClient("mongodb://localhost:27017/")
        db = client["DRS"]
        logger.info(f"Connected to MongoDB successfully | DRS")

    except Exception as err:
        print("conection error")
        logger.error(f"MongoDB connection failed: {str(err)}")
        return False
    
    else:
        try:


            collection = db["Case_Distribution_DRC_Summary"]
            query = {}

            # Check each parameter and build query

            # check drc
            if drc is not None:
                if drc == "D1":
                    query[drc] = {"$regex": f"^{drc}$"}
                elif drc == "D2":
                    query[drc] = drc
                else:
                    raise ValueError(f"Invalid drc '{drc}'. Must be 'D1', or 'D2'")
            

            # check case_distribution_batch_id 
            if case_distribution_batch_id is not None:
                if case_distribution_batch_id == 1:
                    query[case_distribution_batch_id] = {"$regex": f"^{case_distribution_batch_id}$"}
                elif case_distribution_batch_id == 2:
                    query[case_distribution_batch_id] = case_distribution_batch_id
                elif case_distribution_batch_id == 3:
                    query[case_distribution_batch_id] = case_distribution_batch_id
                else:
                    raise ValueError(f"Invalid case distribution batch id '{case_distribution_batch_id}'. Must be 1, 2, or 3")


            #log and excute query
            logger.info(f"Executing query on Case_Distribution_DRC_Summary: {query}")
            summaries = list(collection.find(query)) #fetch data into array
            logger.info(f"Found {len(summaries)} matching DRC summary records")

            if not summaries:
                print("No DRC summary records found matching the selected filters")
                return False

            # Export to Excel even if no incidents are found
            output_dir = "exports"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"drc_summary_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)

            if not create_drc_summary_table(wb, summaries, {
                "drc": drc,
                "case_distribution_batch_id": case_distribution_batch_id
            }):
                raise Exception("Failed to create DRC summary sheet")

            wb.save(filepath)
            if not summaries:
                print("No drc summaries found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(summaries)} DRC summary records to: {filepath}")
            return True

        except ValueError as ve:
            logger.error(f"Validation error: {str(ve)}")
            print(f"Error: {str(ve)}")
            return False
        except Exception as e:
            logger.error(f"Export failed: {str(e)}", exc_info=True)
            print(f"\nError during export: {str(e)}")
            return False
        
        finally:
            if client:
                client.close()
                logger.info("MongoDB connection closed")



def create_drc_summary_table(wb, data, filters=None):
    """Create formatted Excel sheet with DRC summary data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="DRC SUMMARY REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DRC_SUMMARY_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="DRC SUMMARY REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            # Task ID filter
            if filters.get('task_id'):
                ws.cell(row=row_idx, column=2, value="Task ID:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=str(filters['task_id'])).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            # DRC filter
            if filters.get('drc'):
                ws.cell(row=row_idx, column=2, value="DRC:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            # Case Distribution Batch ID filter
            if filters.get('case_distribution_batch_id') is not None:
                ws.cell(row=row_idx, column=2, value="Case Distribution Batch ID:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=str(filters['case_distribution_batch_id'])).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DRC_SUMMARY_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(DRC_SUMMARY_HEADERS, 1):
                    value = record.get(header, "")
                    if header == "drc_id" and isinstance(value, ObjectId):
                        value = str(value)
                    if header == "created_dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    if header == "proceed_on" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(DRC_SUMMARY_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(DRC_SUMMARY_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating DRC summary sheet: {str(e)}", exc_info=True)
        return False