import logging
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
import os
from pymongo import MongoClient


logger = logging.getLogger('excel_data_writer')

CPE_HEADERS = [
    "Incident_Id", "Incident_Status", "Account_Num", "Actions",
    "Created_Dtm"
]


def excel_cpe_detail(from_date, to_date, drc_commision_rule):
    """Fetch and export 'collect CPE' incidents from Incident collection"""

    try:
        client = MongoClient("mongodb://localhost:27017/")
        db = client["DRS"]
        logger.info(f"Connected to MongoDB successfully | DRS")

    except Exception as err:
        print("Connection error")
        logger.error(f"MongoDB connection failed: {str(err)}")
        return False
    else:

        try:
            collection = db["Incident"]
            query = {"Actions": "collect CPE"}  # Fixed to only collect CPE

            # Validate and apply drc_commision_rule filter
            if drc_commision_rule is not None:
                if drc_commision_rule == "PEO TV":
                    query["Drc commision rule"] = {"$regex": f"^{drc_commision_rule}$"}
                elif drc_commision_rule == "BB":
                    query["Actions"] = drc_commision_rule   
                else:
                    raise ValueError(f"Invalid drc_commision_rule '{drc_commision_rule}'. Must be 'PEO TV', 'BB'")
                     

            # Apply date range filter
            if from_date is not None and to_date is not None:
                try:
                    # Check if from_date and to_date are in correct YYYY-MM-DD format
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    # Validate date range
                    if to_dt < from_dt:
                        raise ValueError("to_date cannot be earlier than from_date")
                    
                   # Construct query                  
                    query["Created_Dtm"] = {"$gte": from_dt, "$lte": to_dt}

                except ValueError as ve:
                    if str(ve).startswith("to_date"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")


            logger.info(f"Executing query on Incident for CPE: {query}")
            incidents = list(collection.find(query))
            logger.info(f"Found {len(incidents)} matching CPE incidents")

            # Export to Excel
            output_dir = "exports"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"cpe_incidents_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            os.makedirs(output_dir, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)

            if not create_cpe_table(wb, incidents, {
                "action": "collect CPE",
                "drc_commision_rule": drc_commision_rule,
                "date_range": (datetime.strptime(from_date, '%Y-%m-%d') if from_date else None,
                            datetime.strptime(to_date, '%Y-%m-%d') if to_date else None)
            }):
                raise Exception("Failed to create CPE incident sheet")

            wb.save(filepath)

            if not incidents:
                print("No CPE incidents found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(incidents)} CPE records to: {filepath}")
            return True

        except ValueError as ve:
            logger.error(f"Validation error: {str(ve)}")
            print(f"Error: {str(ve)}")
            return False
        except Exception as e:
            logger.error(f"Export failed: {str(e)}", exc_info=True)
            print(f"\nError during export: {str(e)}")
            return False
        finally:
            if client:
                client.close()
                logger.info("MongoDB connection closed")
    

def create_cpe_table(wb, data, filters=None):
    """Create formatted Excel sheet with CPE incident data"""
    try:
        ws = wb.create_sheet(title="CPE INCIDENT REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(CPE_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="CPE INCIDENT REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            # Action filter (always "collect CPE")
            ws.cell(row=row_idx, column=2, value="Action:").font = STYLES['FilterParam_Style']['font']
            ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
            ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
            ws.cell(row=row_idx, column=3, value=filters['action']).font = STYLES['FilterValue_Style']['font']
            ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
            ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
            row_idx += 1
            
            # DRC Commission Rule filter
            if filters.get('drc_commision_rule'):
                ws.cell(row=row_idx, column=2, value="DRC Commission Rule:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc_commision_rule']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            # Date Range filter
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(CPE_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows
        for record in data:
            row_idx += 1
            for col_idx, header in enumerate(CPE_HEADERS, 1):
                value = record.get(header, "")
                if header == "Incident_Id" and isinstance(value, ObjectId):
                    value = str(value)
                if header == "Created_Dtm" and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = STYLES['Border_Style']['font']
                cell.border = STYLES['Border_Style']['border']
                cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to all columns
        if data:
            last_col_letter = get_column_letter(len(CPE_HEADERS))
            ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{row_idx}"
        
        # Auto-adjust columns
        for col_idx in range(1, len(CPE_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating CPE sheet: {str(e)}", exc_info=True)
        return False