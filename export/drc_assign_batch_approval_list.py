import logging
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
import os
from utils.connectDB import get_db_connection
import logging.config
from utils.config_loader import get_config
from pymongo import MongoClient

logger = logging.getLogger('excel_data_writer')

DRC_ASSIGN_BATCH_APPROVAL_HEADERS = [
    "Batch_id", "created_dtm", "drc_commision_rule", "approval_type", "case_count", "total_arrears"
]

def excel_drc_assign_batch_approval(approver_ref):
    """Fetch and export DRC assign batch approval data based on validated approver_ref parameter"""
    try:
        client = MongoClient("mongodb://localhost:27017/")
        db = client["DRS"]
        logger.info(f"Connected to MongoDB successfully | DRS")

    except Exception as err:
        print("Connection error")
        logger.error(f"MongoDB connection failed: {str(err)}")
        return False
    else:
        try:
            collection = db["Batch_Approval_log"]
            query = {}

            # Check approver_ref
            if approver_ref is not None:
                if approver_ref in ["k1", "k2"]:
                    query["approver_ref"] = approver_ref
                else:
                    raise ValueError("approver_ref must be 'k1' or 'k2'")

            # Log and execute query
            logger.info(f"Executing query: {query}")
            batches = list(collection.find(query))
            logger.info(f"Found {len(batches)} matching batch records")

            # Export to Excel even if no batches are found
            output_dir = "exports"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"drc_assign_batch_approval_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            os.makedirs(output_dir, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)

            if not create_drc_assign_batch_approval_table(wb, batches, {
                "approver_ref": approver_ref
            }):
                raise Exception("Failed to create DRC assign batch approval sheet")

            wb.save(filepath)
            if not batches:
                print(f"No batch approval records found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(batches)} records to: {filepath}")
            return True

        except ValueError as ve:
            logger.error(f"Validation error: {str(ve)}")
            print(f"Error: {str(ve)}")
            return False
        except Exception as e:
            logger.error(f"Export failed: {str(e)}", exc_info=True)
            print(f"\nError during export: {str(e)}")
            return False
        finally:
            if client:
                client.close()
                logger.info("MongoDB connection closed")

def create_drc_assign_batch_approval_table(wb, data, filters=None):
    """Create formatted Excel sheet with DRC assign batch approval data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="DRC ASSIGN BATCH APPROVAL REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DRC_ASSIGN_BATCH_APPROVAL_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="DRC ASSIGN BATCH APPROVAL REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            if filters.get('approver_ref'):
                ws.cell(row=row_idx, column=2, value="Approver Reference:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['approver_ref']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DRC_ASSIGN_BATCH_APPROVAL_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(DRC_ASSIGN_BATCH_APPROVAL_HEADERS, 1):
                    value = record.get(header, "")
                    if header == "Batch_id" and isinstance(value, ObjectId):
                        value = str(value)
                    if header == "created_dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(DRC_ASSIGN_BATCH_APPROVAL_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(DRC_ASSIGN_BATCH_APPROVAL_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False