import logging
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
import os
from pymongo import MongoClient


logger = logging.getLogger('excel_data_writer')

APPROVAL_HEADERS = [
    "case_id", "created_dtm", "created_by", "approval_type",
    "approve_status", "approved_by", "remark"
]

VALID_APPROVAL_TYPES = ["a1", "a2"]

def excel_drc_approval_detail(approval_type, from_date, to_date):
    """Fetch and export DRC assign manager approval details from Case_details collection"""
    try:
        client = MongoClient("mongodb://localhost:27017/")
        db = client["DRS"]
        logger.info(f"Connected to MongoDB successfully | DRS")

    except Exception as err:
        print("Connection error")
        logger.error(f"MongoDB connection failed: {str(err)}")
        return False
    
    else:
        try:
            collection = db["Case_details"]
            query = {}

             # Check date range
            if from_date is not None and to_date is not None:
                try:
                    # Check if from_date and to_date are in correct YYYY-MM-DD format
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    # Validate date range
                    if to_dt < from_dt:
                        raise ValueError("to_date cannot be earlier than from_date")
                    
                   # Construct query                  
                    query["Created_Dtm"] = {"$gte": from_dt, "$lte": to_dt}

                except ValueError as ve:
                    if str(ve).startswith("to_date"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")


            # If approval_type is provided, filter within the approve array
            if approval_type is not None:
                if approval_type == "a1":
                    query["approval_type"] = {"$regex": f"^{approval_type}$"}
                elif approval_type == "a2":
                    query["Incident_Status"] = approval_type
                else:
                    raise ValueError(f"Invalid approval type '{approval_type}'. Must be 'a1', 'a2'")


                    

            logger.info(f"Executing query on Case_details: {query}")
            cases = list(collection.find(query))
            logger.info(f"Found {len(cases)} matching case records")

            if not cases:
                print("No approval records found matching the selected filters")
                return False

            # Process data to flatten the approve array
            processed_data = []
            for case in cases:
                for approval in case.get("approve", []):
                    if not approval_type or approval.get("approval_type") == approval_type:
                        processed_data.append({
                            "case_id": case.get("case_id", ""),
                            "created_dtm": case.get("created_dtm", ""),
                            "created_by": case.get("created_by", ""),
                            "approval_type": approval.get("approval_type", ""),
                            "approve_status": approval.get("approve_status", ""),
                            "approved_by": approval.get("approved_by", ""),
                            "remark": approval.get("remark", "")
                        })

            if not processed_data:
                print("No approval records found within the approve array matching the filters")
                return False

            # Export to Excel
            output_dir= "exports" 
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"drc_approval_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            os.makedirs(output_dir, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)

            if not create_approval_table(wb, processed_data, {
                "approval_type": approval_type,
                "date_range": (datetime.strptime(from_date, '%Y-%m-%d') if from_date else None,
                            datetime.strptime(to_date, '%Y-%m-%d') if to_date else None)
            }):
                raise Exception("Failed to create DRC approval sheet")

            wb.save(filepath)
            print(f"\nSuccessfully exported {len(processed_data)} DRC approval records to: {filepath}")
            return True

        except ValueError as ve:
            logger.error(f"Validation error: {str(ve)}")
            print(f"Error: {str(ve)}")
            return False
        except Exception as e:
            logger.error(f"Export failed: {str(e)}", exc_info=True)
            print(f"\nError during export: {str(e)}")
            return False

def create_approval_table(wb, data, filters=None):
    """Create formatted Excel sheet with DRC approval data"""
    try:
        ws = wb.create_sheet(title="DRC APPROVAL REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(APPROVAL_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="DRC APPROVAL REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            # Approval Type filter
            if filters.get('approval_type'):
                ws.cell(row=row_idx, column=2, value="Approval Type:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['approval_type']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            # Date Range filter
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(APPROVAL_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows
        for record in data:
            row_idx += 1
            for col_idx, header in enumerate(APPROVAL_HEADERS, 1):
                value = record.get(header, "")
                if header == "case_id" and isinstance(value, ObjectId):
                    value = str(value)
                if header == "created_dtm" and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = STYLES['Border_Style']['font']
                cell.border = STYLES['Border_Style']['border']
                cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to all columns
        if data:
            last_col_letter = get_column_letter(len(APPROVAL_HEADERS))
            ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{row_idx}"
        
        # Auto-adjust columns
        for col_idx in range(1, len(APPROVAL_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating DRC approval sheet: {str(e)}", exc_info=True)
        return False