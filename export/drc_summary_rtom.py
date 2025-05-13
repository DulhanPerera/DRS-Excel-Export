import logging
from datetime import datetime
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
import os

logger = logging.getLogger('excel_data_writer')

DRC_SUMMARY_HEADERS = [
    "rtom", "case_count", "tot_arrease"
]

VALID_DRC_VALUES = ["D1", "D2"]

def excel_drc_summary_rtom_detail(db, drc=None, output_path="exports"):
    """Fetch and export DRC summary details from Case_Distribution_DRC_Summary collection"""
    try:
        collection = db["Case_Distribution_DRC_Summary"]
        query = {}

        # Validate and apply drc filter
        if drc and drc.strip():
            if drc not in VALID_DRC_VALUES:
                raise ValueError(f"Invalid drc '{drc}'. Must be one of: {', '.join(VALID_DRC_VALUES)}")
            query["drc"] = drc

        logger.info(f"Executing query on Case_Distribution_DRC_Summary: {query}")
        summaries = list(collection.find(query))
        logger.info(f"Found {len(summaries)} matching DRC summary records")

        if not summaries:
            print("No DRC summary records found matching the selected filters")
            return False

        # Export to Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"drc_summary_rtom_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        if not create_drc_summary_rtom_table(wb, summaries, {"drc": drc}):
            raise Exception("Failed to create DRC summary sheet")

        wb.save(filepath)
        print(f"\nSuccessfully exported {len(summaries)} DRC summary records to: {filepath}")
        return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False

def create_drc_summary_rtom_table(wb, data, filters=None):
    """Create formatted Excel sheet with DRC summary data"""
    try:
        ws = wb.create_sheet(title="DRC SUMMARY REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DRC_SUMMARY_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="DRC SUMMARY REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            # DRC filter
            if filters.get('drc'):
                ws.cell(row=row_idx, column=2, value="DRC:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DRC_SUMMARY_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows
        for record in data:
            row_idx += 1
            for col_idx, header in enumerate(DRC_SUMMARY_HEADERS, 1):
                value = record.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = STYLES['Border_Style']['font']
                cell.border = STYLES['Border_Style']['border']
                cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to all columns
        if data:
            last_col_letter = get_column_letter(len(DRC_SUMMARY_HEADERS))
            ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{row_idx}"
        
        # Auto-adjust columns
        for col_idx in range(1, len(DRC_SUMMARY_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating DRC summary sheet: {str(e)}", exc_info=True)
        return False

