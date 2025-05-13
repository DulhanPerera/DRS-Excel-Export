import logging
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
import os
from utils.connectDB import get_db_connection
import logging.config
from utils.config_loader import get_config
from pymongo import MongoClient

logger = logging.getLogger('excel_data_writer')

INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS = [
    "Id", "Incident_Status", "Account_Num", "Actions",
    "Arrears", "Source_Type"
]

def excel_incident_open_distribution():
    """Fetch and export all open incidents for distribution without parameter filtering"""
    try:
        client = MongoClient("mongodb://localhost:27017/")
        db = client["DRS"]
        logger.info(f"Connected to MongoDB successfully | DRS")

    except Exception as err:
        print("Connection error")
        logger.error(f"MongoDB connection failed: {str(err)}")
        return False
    else:
        try:
            collection = db["Incident_log"]
            query = {"Incident_Status": "Incident Open"}  # Fixed filter for open incidents

            # Log and execute query
            logger.info(f"Executing query: {query}")
            incidents = list(collection.find(query))
            logger.info(f"Found {len(incidents)} matching incidents")

            # Export to Excel even if no incidents are found
            output_dir = "exports"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"incident_open_distribution_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            os.makedirs(output_dir, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)

            if not create_incident_open_distribution_table(wb, incidents):
                raise Exception("Failed to create incident open distribution sheet")

            wb.save(filepath)
            if not incidents:
                print(f"No open incidents found. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(incidents)} records to: {filepath}")
            return True

        except Exception as e:
            logger.error(f"Export failed: {str(e)}", exc_info=True)
            print(f"\nError during export: {str(e)}")
            return False
        finally:
            if client:
                client.close()
                logger.info("MongoDB connection closed")

def create_incident_open_distribution_table(wb, data):
    """Create formatted Excel sheet with open incident distribution data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="OPEN INCIDENT DISTRIBUTION")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="OPEN INCIDENT DISTRIBUTION REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 2
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS, 1):
                    value = record.get(header, "")
                    if header == "Id" and isinstance(value, ObjectId):
                        value = str(value)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False