import configparser
import logging
import logging.config
import os
import sys
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from bson import ObjectId

# Load logger configuration
logging.config.fileConfig('Config/logger/loggers.ini')
logger = logging.getLogger('excel_data_writer')

def get_config():
    """
    Load and return the configuration from the Config.ini file.
    """
    try:
        config = configparser.ConfigParser()
        config.read(os.path.join(os.path.dirname(__file__), '../Config/Config.ini'))
        if not config.sections():
            logger.error("Configuration file is empty or not found.")
            sys.exit(1)
        return config
    except Exception as e:
        logger.error(f"Failed to load configuration: {e}")
        sys.exit(1)

def connect_db(config):
    """
    Connect to the MongoDB database using the configuration.
    """
    try:
        client = MongoClient(config['DATABASE']['MONGO_URI'])
        db = client[config['DATABASE']['DB_NAME']]
        logger.info("Successfully connected to the database.")
        return db
    except Exception as e:
        logger.error(f"Failed to connect to the database: {e}")
        sys.exit(1)

def format_with_thousand_separator(value):
    """
    Format numeric values with thousand separators.
    """
    if isinstance(value, (int, float)):
        return f"{value:,}"  # Add thousand separators
    return value

def get_arrears_band_value(db, current_arrears_band):
    """
    Retrieve the value for the given arrears band from the arrears_bands collection.
    """
    try:
        arrears_bands_collection = db["Arrears_bands"]
        arrears_bands_doc = arrears_bands_collection.find_one({})
        if arrears_bands_doc:
            return arrears_bands_doc.get(current_arrears_band)
        else:
            logger.warning("No arrears bands document found in the collection.")
            return None
    except Exception as e:
        logger.error(f"Failed to retrieve arrears band value: {e}")
        return None

def get_settlement_data(db, case_id):
    """
    Retrieve settlement data for the given case_id from the case_settlements collection.
    """
    try:
        settlements_collection = db["Case_settlements"]
        settlements = list(settlements_collection.find({"case_id": case_id}))
        if settlements:
            logger.info(f"Found {len(settlements)} settlement records for case_id: {case_id}")
        else:
            logger.warning(f"No settlement records found for case_id: {case_id}")
        return settlements
    except Exception as e:
        logger.error(f"Failed to retrieve settlement data: {e}")
        return []

def get_settlement_plan_data(db, case_id):
    """
    Retrieve settlement plan data for the given case_id from the case_settlements collection.
    """
    try:
        settlements_collection = db["Case_settlements"]
        settlements = list(settlements_collection.find({"case_id": case_id}))
        settlement_plans = []
        for settlement in settlements:
            if "settlement_plan" in settlement:
                for plan in settlement["settlement_plan"]:
                    # Add settlement_id to each plan
                    plan["settlement_id"] = settlement.get("settlement_id")
                    settlement_plans.append(plan)
        if settlement_plans:
            logger.info(f"Found {len(settlement_plans)} settlement plan records for case_id: {case_id}")
        else:
            logger.warning(f"No settlement plan records found for case_id: {case_id}")
        return settlement_plans
    except Exception as e:
        logger.error(f"Failed to retrieve settlement plan data: {e}")
        return []

def create_case_details_table(ws, case_data, start_row, start_col, db):
    """
    Create the Case Details table in the worksheet.
    """
    try:
        logger.info("Creating Case Details table...")
        headers = [
            "Case ID", "Incident ID", "Account No.", "Customer Ref", "Area",
            "BSS Arrears Amount", "Current Arrears Amount", "Action type", "Filtered reason",
            "Last Payment Date", "Last BSS Reading Date", "Commission", "Case Current Status",
            "Current Arrears band", "DRC Commission Rule", "Created dtm", "Implemented dtm",
            "RTOM", "Monitor months"
        ]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        bold_font = Font(bold=True)
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        ws.cell(row=start_row, column=start_col, value="Case details").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=start_row + idx, column=start_col, value=header).font = header_font
            ws.cell(row=start_row + idx, column=start_col).fill = header_fill
            ws.cell(row=start_row + idx, column=start_col).border = cell_border
            ws.cell(row=start_row + idx, column=start_col).alignment = header_alignment
        
        # Map MongoDB data to headers
        data_mapping = {
            "Case ID": case_data.get("case_id"),
            "Incident ID": case_data.get("incident_id"),
            "Account No.": case_data.get("account_no"),
            "Customer Ref": case_data.get("customer_ref"),
            "Area": case_data.get("area"),
            "BSS Arrears Amount": format_with_thousand_separator(case_data.get("bss_arrears_amount")),
            "Current Arrears Amount": format_with_thousand_separator(case_data.get("current_arrears_amount")),
            "Action type": case_data.get("action_type"),
            "Filtered reason": case_data.get("filtered_reason"),
            "Last Payment Date": case_data.get("last_payment_date"),
            "Last BSS Reading Date": case_data.get("last_bss_reading_date"),
            "Commission": format_with_thousand_separator(case_data.get("commission")),
            "Case Current Status": case_data.get("case_current_status"),
            "Current Arrears band": case_data.get("current_arrears_band"),
            "DRC Commission Rule": case_data.get("drc_commision_rule"),
            "Created dtm": case_data.get("created_dtm"),
            "Implemented dtm": case_data.get("implemented_dtm"),
            "RTOM": case_data.get("rtom"),
            "Monitor months": case_data.get("monitor_months")
        }
        
        # Retrieve arrears band value
        current_arrears_band = case_data.get("current_arrears_band")
        if current_arrears_band:
            arrears_band_value = get_arrears_band_value(db, current_arrears_band)
            if arrears_band_value:
                data_mapping["Current Arrears band"] = arrears_band_value
            else:
                logger.warning(f"No value found for arrears band: {current_arrears_band}")
        
        # Insert data
        for idx, header in enumerate(headers, start=1):
            value = data_mapping.get(header)
            if isinstance(value, (list, dict)):
                value = str(value)
            cell = ws.cell(row=start_row + idx, column=start_col + 1, value=value)
            cell.border = cell_border
            if header in ["Case ID", "Incident ID"]:
                cell.font = bold_font
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        logger.info("Case Details table created successfully.")
        return start_row + len(headers) + 1
    except Exception as e:
        logger.error(f"Failed to create Case Details table: {e}")
        sys.exit(1)

def create_contact_details_table(ws, case_data, start_row, start_col):
    """
    Create the Contact Details table in the worksheet.
    """
    try:
        logger.info("Creating Contact Details table...")
        contacts_headers = ["Mobile", "Email", "Home Phone", "Address"]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
        ws.cell(row=start_row, column=start_col, value="Contact Info").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(contacts_headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert contact data
        contacts = case_data.get("contact", [])
        for contact_idx, contact in enumerate(contacts, start=1):
            ws.cell(row=start_row + 1 + contact_idx, column=start_col, value=contact.get("mob")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 1, value=contact.get("email")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 2, value=contact.get("lan")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 3, value=contact.get("address")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(contacts_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(contacts) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Contact Details table created successfully.")
        return start_row + len(contacts) + 3
    except Exception as e:
        logger.error(f"Failed to create Contact Details table: {e}")
        sys.exit(1)

def create_remarks_table(ws, case_data, start_row, start_col):
    """
    Create the Remarks table in the worksheet.
    """
    try:
        logger.info("Creating Remarks table...")
        headers = ["Remark", "Remark Added by", "Remark Added Date"]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
        ws.cell(row=start_row, column=start_col, value="Remarks").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert remarks data
        remarks = case_data.get("remark", [])
        for remark_idx, remark in enumerate(remarks, start=1):
            ws.cell(row=start_row + 1 + remark_idx, column=start_col, value=remark.get("remark")).border = cell_border
            ws.cell(row=start_row + 1 + remark_idx, column=start_col + 1, value=remark.get("remark_added_by")).border = cell_border
            ws.cell(row=start_row + 1 + remark_idx, column=start_col + 2, value=remark.get("remark_added_date")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(remarks) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Remarks table created successfully.")
        return start_row + len(remarks) + 3
    except Exception as e:
        logger.error(f"Failed to create Remarks table: {e}")
        sys.exit(1)

def create_settlement_table(ws, settlements, start_row, start_col):
    """
    Create the Settlement table in the worksheet.
    """
    try:
        logger.info("Creating Settlement table...")
        headers = [
            "Settlement ID", "Case ID", "DRC Name", "RO Name", "Status", "Status reason",
            "Status DTM", "Settlement Type", "Settlement Amount", "Settlement Phase",
            "Settlement Created by", "Settlement Created DTM", "Last Monitoring DTM", "Remark"
        ]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(headers) - 1)
        ws.cell(row=start_row, column=start_col, value="Settlement Details").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert settlement data
        for settlement_idx, settlement in enumerate(settlements, start=1):
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col, value=settlement.get("settlement_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 1, value=settlement.get("case_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 2, value=settlement.get("drc_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 3, value=settlement.get("ro_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 4, value=settlement.get("settlement_status")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 5, value=settlement.get("status_reason")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 6, value=settlement.get("status_dtm")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 7, value=settlement.get("settlement_type")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 8, value=settlement.get("settlement_amount")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 9, value=settlement.get("settlement_phase")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 10, value=settlement.get("created_by")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 11, value=settlement.get("created_on")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 12, value=settlement.get("last_monitoring_dtm")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 13, value=settlement.get("remark")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(settlements) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Settlement table created successfully.")
        return start_row + len(settlements) + 3
    except Exception as e:
        logger.error(f"Failed to create Settlement table: {e}")
        sys.exit(1)

def create_settlement_plan_table(ws, settlement_plans, start_row, start_col):
    """
    Create the Settlement Plan table in the worksheet.
    """
    try:
        logger.info("Creating Settlement Plan table...")
        headers = [
            "Settlement ID", "Installment Sequence", "Installment Settle Amount",
            "Accumulated Amount", "Plan Date and Time"
        ]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(headers) - 1)
        ws.cell(row=start_row, column=start_col, value="Settlement Plan").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert settlement plan data
        for plan_idx, plan in enumerate(settlement_plans, start=1):
            ws.cell(row=start_row + 1 + plan_idx, column=start_col, value=plan.get("settlement_id")).border = cell_border
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 1, value=plan.get("installment_seq")).border = cell_border
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 2, value=plan.get("installment_settle_amount")).border = cell_border
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 3, value=plan.get("accumulated_amount")).border = cell_border
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 4, value=plan.get("plan_date")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(settlement_plans) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Settlement Plan table created successfully.")
        return start_row + len(settlement_plans) + 3
    except Exception as e:
        logger.error(f"Failed to create Settlement Plan table: {e}")
        sys.exit(1)

def create_case_details_sheet(wb, case_data, db):
    """
    Create the Case Details sheet with all tables.
    """
    try:
        logger.info("Creating Case Details sheet...")
        ws = wb.active
        ws.title = "Case Details"
        
        # Define starting row and column for the first table
        start_row, start_col = 2, 1
        
        # Create the Case Details table
        next_row = create_case_details_table(ws, case_data, start_row, start_col, db)
        
        # Add a two-row gap between the tables
        gap_row = next_row + 2
        
        # Create the Contact Details table
        next_row = create_contact_details_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Contact Info and Remarks tables
        gap_row = next_row + 1
        
        # Create the Remarks table
        next_row = create_remarks_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Remarks and Settlement tables
        gap_row = next_row + 1
        
        # Retrieve settlement data for the case
        case_id = case_data.get("case_id")
        settlements = get_settlement_data(db, case_id)
        
        # Create the Settlement table if settlement data exists
        if settlements:
            next_row = create_settlement_table(ws, settlements, gap_row, start_col)
            gap_row = next_row + 1  # Add a gap after the Settlement table
        
        # Retrieve settlement plan data for the case
        settlement_plans = get_settlement_plan_data(db, case_id)
        
        # Create the Settlement Plan table if settlement plan data exists
        if settlement_plans:
            create_settlement_plan_table(ws, settlement_plans, gap_row, start_col)
        
        logger.info("Case Details sheet created successfully.")
        return ws
    except Exception as e:
        logger.error(f"Failed to create Case Details sheet: {e}")
        sys.exit(1)

def export_case_details(db, incident_id, output_path, collection_name):
    """
    Export case details to an Excel file.
    """
    try:
        logger.info(f"Exporting case details for Incident ID: {incident_id}")
        case_collection = db[collection_name]
        case_data = case_collection.find_one({"incident_id": incident_id})
        
        if not case_data:
            logger.error(f"No case details found for Incident ID: {incident_id}")
            sys.exit(1)
        
        logger.info(f"Case data found: {case_data}")
        
        wb = Workbook()
        ws = create_case_details_sheet(wb, case_data, db)
        
        if not output_path.endswith('.xlsx'):
            output_path = os.path.join(output_path, 'case_details.xlsx')
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        try:
            wb.save(output_path)
            logger.info(f"Case details exported to {output_path}")
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            sys.exit(1)
    except Exception as e:
        logger.error(f"Failed to export case details: {e}")
        sys.exit(1)

def main():
    """
    Main function to execute the case details export process.
    """
    try:
        logger.info("Starting case details export process...")
        config = get_config()
        db = connect_db(config)
        incident_id = 78910  # Replace with the actual incident ID you want to export
        export_path = config['EXCEL_EXPORT_PATHS']['WIN_DB']
        collection_name = config['COLLECTIONS']['CASE_DETAIL_COLLECTION_NAME']
        
        export_case_details(db, incident_id, export_path, collection_name)
        
        logger.info("Case details export process completed.")
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()