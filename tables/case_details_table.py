from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.data_formatter import format_with_thousand_separator
from manipulation.data_fetcher import get_arrears_band_value
import logging
import sys

logger = logging.getLogger('excel_data_writer')

def create_case_details_table(ws, case_data, start_row, start_col, db):
    """
    Create the Case Details table in the worksheet.
    """
    try:
        logger.info("Creating Case Details table...")
        
        # Define headers for the Case Details table
        case_details_headers = [
            "Case ID", "Incident ID", "Account No.", "Customer Ref", "Area",
            "BSS Arrears Amount", "Current Arrears Amount", "Action type", "Filtered reason",
            "Last Payment Date", "Last BSS Reading Date", "Commission", "Case Current Status",
            "Current Arrears band", "DRC Commission Rule", "Created dtm", "Implemented dtm",
            "RTOM", "Monitor months"
        ]
        
        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Case Details")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers
        for idx, header in enumerate(case_details_headers, start=1):
            header_cell = ws.cell(row=start_row + idx, column=start_col, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        # Map MongoDB data to headers
        data_mapping = {
            "Case ID": case_data.get("case_id"),
            "Incident ID": case_data.get("incident_id"),
            "Account No.": case_data.get("account_no"),
            "Customer Ref": case_data.get("customer_ref"),
            "Area": case_data.get("area"),
            "BSS Arrears Amount": format_with_thousand_separator(case_data.get("bss_arrears_amount")),
            "Current Arrears Amount": format_with_thousand_separator(case_data.get("current_arrears_amount")),
            "Action type": case_data.get("action_type"),
            "Filtered reason": case_data.get("filtered_reason"),
            "Last Payment Date": case_data.get("last_payment_date"),
            "Last BSS Reading Date": case_data.get("last_bss_reading_date"),
            "Commission": format_with_thousand_separator(case_data.get("commission")),
            "Case Current Status": case_data.get("case_current_status"),
            "Current Arrears band": case_data.get("current_arrears_band"),
            "DRC Commission Rule": case_data.get("drc_commision_rule"),
            "Created dtm": case_data.get("created_dtm"),
            "Implemented dtm": case_data.get("implemented_dtm"),
            "RTOM": case_data.get("rtom"),
            "Monitor months": case_data.get("monitor_months")
        }
        
        # Retrieve arrears band value (readable format)
        current_arrears_band = case_data.get("current_arrears_band")
        if current_arrears_band:
            arrears_band_value = get_arrears_band_value(db, current_arrears_band)
            if arrears_band_value:
                data_mapping["Current Arrears band"] = arrears_band_value
            else:
                logger.warning(f"No value found for arrears band: {current_arrears_band}")
        
        # Insert data into the table
        for idx, header in enumerate(case_details_headers, start=1):
            value = data_mapping.get(header)
            if isinstance(value, (list, dict)):
                value = str(value)  # Convert lists or dicts to strings
            cell = ws.cell(row=start_row + idx, column=start_col + 1, value=value)
            cell.border = border_style['border']
            if header in ["Case ID", "Incident ID"]:
                cell.font = Font(bold=True)  # Make Case ID and Incident ID bold
        
        # Adjust column widths for the current table
        for col_idx in range(start_col, start_col + 2):  # Only adjust columns for the current table
            max_length = 0
            column_letter = chr(64 + col_idx)
            for row in range(start_row, start_row + len(case_details_headers) + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Case Details table created successfully.")
        return start_row + len(case_details_headers) + 1  # Return the next row after the table
    except Exception as e:
        logger.error(f"Failed to create Case Details table: {e}")
        sys.exit(1)