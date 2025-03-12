# case_details_table.py
# import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from data_fetcher import get_arrears_band_value
import logging

def format_with_thousand_separator(value):
    """
    Format numeric values with thousand separators.
    """
    if isinstance(value, (int, float)):
        return f"{value:,}"  # Add thousand separators
    return value

def create_case_details_table(ws, case_data, start_row, start_col, db):
    """
    Create the Case Details table in the worksheet.
    """
    try:
        logging.info("Creating Case Details table...")
        headers = [
            "Case ID", "Incident ID", "Account No.", "Customer Ref", "Area",
            "BSS Arrears Amount", "Current Arrears Amount", "Action type", "Filtered reason",
            "Last Payment Date", "Last BSS Reading Date", "Commission", "Case Current Status",
            "Current Arrears band", "DRC Commission Rule", "Created dtm", "Implemented dtm",
            "RTOM", "Monitor months"
        ]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        bold_font = Font(bold=True)
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        ws.cell(row=start_row, column=start_col, value="Case details").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=start_row + idx, column=start_col, value=header).font = header_font
            ws.cell(row=start_row + idx, column=start_col).fill = header_fill
            ws.cell(row=start_row + idx, column=start_col).border = cell_border
            ws.cell(row=start_row + idx, column=start_col).alignment = header_alignment
        
        # Map MongoDB data to headers
        data_mapping = {
            "Case ID": case_data.get("case_id"),
            "Incident ID": case_data.get("incident_id"),
            "Account No.": case_data.get("account_no"),
            "Customer Ref": case_data.get("customer_ref"),
            "Area": case_data.get("area"),
            "BSS Arrears Amount": format_with_thousand_separator(case_data.get("bss_arrears_amount")),
            "Current Arrears Amount": format_with_thousand_separator(case_data.get("current_arrears_amount")),
            "Action type": case_data.get("action_type"),
            "Filtered reason": case_data.get("filtered_reason"),
            "Last Payment Date": case_data.get("last_payment_date"),
            "Last BSS Reading Date": case_data.get("last_bss_reading_date"),
            "Commission": format_with_thousand_separator(case_data.get("commission")),
            "Case Current Status": case_data.get("case_current_status"),
            "Current Arrears band": case_data.get("current_arrears_band"),
            "DRC Commission Rule": case_data.get("drc_commision_rule"),
            "Created dtm": case_data.get("created_dtm"),
            "Implemented dtm": case_data.get("implemented_dtm"),
            "RTOM": case_data.get("rtom"),
            "Monitor months": case_data.get("monitor_months")
        }
        
        # Retrieve arrears band value
        current_arrears_band = case_data.get("current_arrears_band")
        if current_arrears_band:
            arrears_band_value = get_arrears_band_value(db, current_arrears_band)
            if arrears_band_value:
                data_mapping["Current Arrears band"] = arrears_band_value
            else:
                logging.warning(f"No value found for arrears band: {current_arrears_band}")
        
        # Insert data
        for idx, header in enumerate(headers, start=1):
            value = data_mapping.get(header)
            if isinstance(value, (list, dict)):
                value = str(value)
            cell = ws.cell(row=start_row + idx, column=start_col + 1, value=value)
            cell.border = cell_border
            if header in ["Case ID", "Incident ID"]:
                cell.font = bold_font
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        logging.info("Case Details table created successfully.")
        return start_row + len(headers) + 1
    except Exception as e:
        logging.error(f"Failed to create Case Details table: {e}")
        sys.exit(1)