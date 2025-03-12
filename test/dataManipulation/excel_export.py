# excel_export.py
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import logging
import os
from data_retrieval import get_arrears_band_value, get_settlement_data, get_settlement_plan_data

# Rest of the code remains the same

def format_with_thousand_separator(value):
    """
    Format numeric values with thousand separators.
    """
    if isinstance(value, (int, float)):
        return f"{value:,}"  # Add thousand separators
    return value

def create_case_details_table(ws, case_data, start_row, start_col, db):
    """
    Create the Case Details table in the worksheet.
    """
    try:
        logging.info("Creating Case Details table...")
        headers = [
            "Case ID", "Incident ID", "Account No.", "Customer Ref", "Area",
            "BSS Arrears Amount", "Current Arrears Amount", "Action type", "Filtered reason",
            "Last Payment Date", "Last BSS Reading Date", "Commission", "Case Current Status",
            "Current Arrears band", "DRC Commission Rule", "Created dtm", "Implemented dtm",
            "RTOM", "Monitor months"
        ]
        # ... (rest of the function remains the same)
        logging.info("Case Details table created successfully.")
        return start_row + len(headers) + 1
    except Exception as e:
        logging.error(f"Failed to create Case Details table: {e}")
        sys.exit(1)

def create_contact_details_table(ws, case_data, start_row, start_col):
    """
    Create the Contact Details table in the worksheet.
    """
    try:
        logging.info("Creating Contact Details table...")
        contacts_headers = ["Mobile", "Email", "Home Phone", "Address"]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
        ws.cell(row=start_row, column=start_col, value="Contact Info").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(contacts_headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert contact data
        contacts = case_data.get("contact", [])
        for contact_idx, contact in enumerate(contacts, start=1):
            ws.cell(row=start_row + 1 + contact_idx, column=start_col, value=contact.get("mob")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 1, value=contact.get("email")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 2, value=contact.get("lan")).border = cell_border
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 3, value=contact.get("address")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(contacts_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(contacts) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logging.info("Contact Details table created successfully.")
        return start_row + len(contacts) + 3
    except Exception as e:
        logging.error(f"Failed to create Contact Details table: {e}")
        sys.exit(1)

def create_case_details_sheet(wb, case_data, db):
    """
    Create the Case Details sheet with all tables.
    """
    try:
        logging.info("Creating Case Details sheet...")
        ws = wb.active
        ws.title = "Case Details"
        
        # Define starting row and column for the first table
        start_row, start_col = 2, 1
        
        # Create the Case Details table
        next_row = create_case_details_table(ws, case_data, start_row, start_col, db)
        
        # Add a two-row gap between the tables
        gap_row = next_row + 2
        
        # Create the Contact Details table
        next_row = create_contact_details_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Contact Info and Remarks tables
        gap_row = next_row + 1
        
        # Create the Remarks table
        next_row = create_remarks_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Remarks and Settlement tables
        gap_row = next_row + 1
        
        # Retrieve settlement data for the case
        case_id = case_data.get("case_id")
        settlements = get_settlement_data(db, case_id)
        
        # Create the Settlement table if settlement data exists
        if settlements:
            next_row = create_settlement_table(ws, settlements, gap_row, start_col)
            gap_row = next_row + 1  # Add a gap after the Settlement table
        
        # Retrieve settlement plan data for the case
        settlement_plans = get_settlement_plan_data(db, case_id)
        
        # Create the Settlement Plan table if settlement plan data exists
        if settlement_plans:
            create_settlement_plan_table(ws, settlement_plans, gap_row, start_col)
        
        logging.info("Case Details sheet created successfully.")
        return ws
    except Exception as e:
        logging.error(f"Failed to create Case Details sheet: {e}")
        sys.exit(1)

def export_case_details(db, incident_id, output_path, collection_name):
    """
    Export case details to an Excel file.
    """
    try:
        logging.info(f"Exporting case details for Incident ID: {incident_id}")
        case_collection = db[collection_name]
        case_data = case_collection.find_one({"incident_id": incident_id})
        
        if not case_data:
            logging.error(f"No case details found for Incident ID: {incident_id}")
            sys.exit(1)
        
        logging.info(f"Case data found: {case_data}")
        
        wb = Workbook()
        ws = create_case_details_sheet(wb, case_data, db)
        
        if not output_path.endswith('.xlsx'):
            output_path = os.path.join(output_path, 'case_details.xlsx')
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        try:
            wb.save(output_path)
            logging.info(f"Case details exported to {output_path}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")
            sys.exit(1)
    except Exception as e:
        logging.error(f"Failed to export case details: {e}")
        sys.exit(1)