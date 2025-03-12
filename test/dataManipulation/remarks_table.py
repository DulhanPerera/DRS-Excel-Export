# remarks_table.py
import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import logging

def create_remarks_table(ws, case_data, start_row, start_col):
    """
    Create the Remarks table in the worksheet.
    """
    try:
        logging.info("Creating Remarks table...")
        headers = ["Remark", "Remark Added by", "Remark Added Date"]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
        ws.cell(row=start_row, column=start_col, value="Remarks").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert remarks data
        remarks = case_data.get("remark", [])
        for remark_idx, remark in enumerate(remarks, start=1):
            ws.cell(row=start_row + 1 + remark_idx, column=start_col, value=remark.get("remark")).border = cell_border
            ws.cell(row=start_row + 1 + remark_idx, column=start_col + 1, value=remark.get("remark_added_by")).border = cell_border
            ws.cell(row=start_row + 1 + remark_idx, column=start_col + 2, value=remark.get("remark_added_date")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(remarks) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logging.info("Remarks table created successfully.")
        return start_row + len(remarks) + 3
    except Exception as e:
        logging.error(f"Failed to create Remarks table: {e}")
        sys.exit(1)