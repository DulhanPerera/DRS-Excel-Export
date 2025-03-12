# settlement_table.py
import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import logging

def create_settlement_table(ws, settlements, start_row, start_col):
    """
    Create the Settlement table in the worksheet.
    """
    try:
        logging.info("Creating Settlement table...")
        headers = [
            "Settlement ID", "Case ID", "DRC Name", "RO Name", "Status", "Status reason",
            "Status DTM", "Settlement Type", "Settlement Amount", "Settlement Phase",
            "Settlement Created by", "Settlement Created DTM", "Last Monitoring DTM", "Remark"
        ]
        header_font = Font(bold=True, color="000000", size=12)
        main_header_fill = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_alignment = Alignment(horizontal="left", vertical="center")
        main_header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(headers) - 1)
        ws.cell(row=start_row, column=start_col, value="Settlement Details").font = Font(bold=True, color="000000", size=16)
        ws.cell(row=start_row, column=start_col).fill = main_header_fill
        ws.cell(row=start_row, column=start_col).border = cell_border
        ws.cell(row=start_row, column=start_col).alignment = main_header_alignment
        
        # Write headers
        for idx, header in enumerate(headers, start=0):
            ws.cell(row=start_row + 1, column=start_col + idx, value=header).font = header_font
            ws.cell(row=start_row + 1, column=start_col + idx).fill = header_fill
            ws.cell(row=start_row + 1, column=start_col + idx).border = cell_border
            ws.cell(row=start_row + 1, column=start_col + idx).alignment = header_alignment
        
        # Insert settlement data
        for settlement_idx, settlement in enumerate(settlements, start=1):
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col, value=settlement.get("settlement_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 1, value=settlement.get("case_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 2, value=settlement.get("drc_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 3, value=settlement.get("ro_id")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 4, value=settlement.get("settlement_status")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 5, value=settlement.get("status_reason")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 6, value=settlement.get("status_dtm")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 7, value=settlement.get("settlement_type")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 8, value=settlement.get("settlement_amount")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 9, value=settlement.get("settlement_phase")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 10, value=settlement.get("created_by")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 11, value=settlement.get("created_on")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 12, value=settlement.get("last_monitoring_dtm")).border = cell_border
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 13, value=settlement.get("remark")).border = cell_border
        
        # Adjust column widths
        for col in range(start_col, start_col + len(headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(settlements) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logging.info("Settlement table created successfully.")
        return start_row + len(settlements) + 3
    except Exception as e:
        logging.error(f"Failed to create Settlement table: {e}")
        sys.exit(1)