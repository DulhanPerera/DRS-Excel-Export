import configparser
import logging
import logging.config
import os
import sys
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from bson import ObjectId
import datetime


# Load logger configuration
logging.config.fileConfig('Config/logger/loggers.ini')
logger = logging.getLogger('excel_data_writer')


# Global dictionary to store styles
STYLES = {}

def load_styles_from_ini(ini_file_path):
    """
    Load styles from the INI file and store them in the global STYLES dictionary.
    """
    global STYLES
    config = configparser.ConfigParser()
    try:
        config.read(ini_file_path)
        if not config.sections():
            logger.error(f"INI file '{ini_file_path}' is empty or not found.")
            sys.exit(1)

        for section in config.sections():
            STYLES[section] = {}
            for key, value in config[section].items():
                if key == 'font':
                    font_attrs = {}
                    for attr in value.split(','):
                        attr_name, attr_value = attr.strip().split('=')
                        # Convert attribute values to appropriate types
                        if attr_name == "size":
                            attr_value = int(attr_value)
                        elif attr_name in ["bold", "italic"]:
                            attr_value = attr_value.lower() == "true"
                        elif attr_name == "color":
                            # Convert color value to a Color object
                            if not attr_value.startswith("FF"):
                                attr_value = "FF" + attr_value  # Ensure alpha channel
                            attr_value = Color(rgb=attr_value)
                        # Only include valid Font attributes
                        if attr_name in ["name", "size", "bold", "italic", "color"]:
                            font_attrs[attr_name] = attr_value
                    STYLES[section]['font'] = Font(**font_attrs)
                elif key == 'fill':
                    fill_attrs = {}
                    for attr in value.split(','):
                        attr_name, attr_value = attr.strip().split('=')
                        if attr_name in ["start_color", "end_color"]:
                            # Convert color value to a Color object
                            if not attr_value.startswith("FF"):
                                attr_value = "FF" + attr_value  # Ensure alpha channel
                            attr_value = Color(rgb=attr_value)
                        fill_attrs[attr_name] = attr_value
                    STYLES[section]['fill'] = PatternFill(**fill_attrs)
                elif key == 'alignment':
                    alignment_attrs = {}
                    for attr in value.split(','):
                        attr_name, attr_value = attr.strip().split('=')
                        alignment_attrs[attr_name] = attr_value
                    STYLES[section]['alignment'] = Alignment(**alignment_attrs)
                elif key == 'border':
                    border_attrs = {}
                    for attr in value.split(','):
                        attr_name, attr_value = attr.strip().split('=')
                        border_attrs[attr_name] = Side(style=attr_value)
                    STYLES[section]['border'] = Border(**border_attrs)
                elif key == 'width':
                    STYLES[section]['width'] = float(value)
    except Exception as e:
        logger.error(f"Failed to load styles from INI file: {e}")
        sys.exit(1)

# Load styles from the INI file
load_styles_from_ini('Config/table_format.ini')


def get_config():
    """
    Load and return the configuration from the Config.ini file.
    """
    try:
        config = configparser.ConfigParser()
        config.read(os.path.join(os.path.dirname(__file__), '../Config/Config.ini'))
        if not config.sections():
            logger.error("Configuration file is empty or not found.")
            sys.exit(1)
        return config
    except Exception as e:
        logger.error(f"Failed to load configuration: {e}")
        sys.exit(1)

def connect_db(config):
    """
    Connect to the MongoDB database using the configuration.
    """
    try:
        client = MongoClient(config['DATABASE']['MONGO_URI'])
        db = client[config['DATABASE']['DB_NAME']]
        logger.info("Successfully connected to the database.")
        return db
    except Exception as e:
        logger.error(f"Failed to connect to the database: {e}")
        sys.exit(1)

def format_with_thousand_separator(value):
    """
    Format numeric values with thousand separators.
    """
    if isinstance(value, (int, float)):
        return f"{value:,}"  # Add thousand separators
    return value

def get_arrears_band_value(db, current_arrears_band):
    """
    Retrieve the value for the given arrears band from the arrears_bands collection.
    """
    try:
        arrears_bands_collection = db["Arrears_bands"]
        arrears_bands_doc = arrears_bands_collection.find_one({})
        if arrears_bands_doc:
            return arrears_bands_doc.get(current_arrears_band)
    except Exception as e:
        logger.error(f"Failed to retrieve arrears band value: {e}")
        return None

def get_settlement_data(db, case_id):
    """
    Retrieve settlement data for the given case_id from the case_settlements collection.
    """
    try:
        settlements_collection = db["Case_settlements"]
        settlements = list(settlements_collection.find({"case_id": case_id}))
        if settlements:
            logger.info(f"Found {len(settlements)} settlement records for case_id: {case_id}")
        return settlements
    except Exception as e:
        logger.error(f"Failed to retrieve settlement data: {e}")
        return []

def get_settlement_plan_data(db, case_id):
    """
    Retrieve settlement plan data for the given case_id from the case_settlements collection.
    """
    try:
        settlements_collection = db["Case_settlements"]
        settlements = list(settlements_collection.find({"case_id": case_id}))
        settlement_plans = []
        for settlement in settlements:
            if "settlement_plan" in settlement:
                for plan in settlement["settlement_plan"]:
                    # Add settlement_id to each plan
                    plan["settlement_id"] = settlement.get("settlement_id")
                    settlement_plans.append(plan)
        if settlement_plans:
            logger.info(f"Found {len(settlement_plans)} settlement plan records for case_id: {case_id}")
        return settlement_plans
    except Exception as e:
        logger.error(f"Failed to retrieve settlement plan data: {e}")
        return []

def create_case_details_table(ws, case_data, start_row, start_col, db):
    """
    Create the Case Details table in the worksheet using styles from the INI file.
    """
    try:
        logger.info("Creating Case Details table...")
        
        # Define headers for the Case Details table
        case_details_headers = [
            "Case ID", "Incident ID", "Account No.", "Customer Ref", "Area",
            "BSS Arrears Amount", "Current Arrears Amount", "Action type", "Filtered reason",
            "Last Payment Date", "Last BSS Reading Date", "Commission", "Case Current Status",
            "Current Arrears band", "DRC Commission Rule", "Created dtm", "Implemented dtm",
            "RTOM", "Monitor months"
        ]
        
        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']
        
        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Case Details")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers
        for idx, header in enumerate(case_details_headers, start=1):
            header_cell = ws.cell(row=start_row + idx, column=start_col, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        # Map MongoDB data to headers
        data_mapping = {
            "Case ID": case_data.get("case_id"),
            "Incident ID": case_data.get("incident_id"),
            "Account No.": case_data.get("account_no"),
            "Customer Ref": case_data.get("customer_ref"),
            "Area": case_data.get("area"),
            "BSS Arrears Amount": format_with_thousand_separator(case_data.get("bss_arrears_amount")),
            "Current Arrears Amount": format_with_thousand_separator(case_data.get("current_arrears_amount")),
            "Action type": case_data.get("action_type"),
            "Filtered reason": case_data.get("filtered_reason"),
            "Last Payment Date": case_data.get("last_payment_date"),
            "Last BSS Reading Date": case_data.get("last_bss_reading_date"),
            "Commission": format_with_thousand_separator(case_data.get("commission")),
            "Case Current Status": case_data.get("case_current_status"),
            "Current Arrears band": case_data.get("current_arrears_band"),
            "DRC Commission Rule": case_data.get("drc_commision_rule"),
            "Created dtm": case_data.get("created_dtm"),
            "Implemented dtm": case_data.get("implemented_dtm"),
            "RTOM": case_data.get("rtom"),
            "Monitor months": case_data.get("monitor_months")
        }
        
        # Retrieve arrears band value (readable format)
        current_arrears_band = case_data.get("current_arrears_band")
        if current_arrears_band:
            arrears_band_value = get_arrears_band_value(db, current_arrears_band)
            if arrears_band_value:
                data_mapping["Current Arrears band"] = arrears_band_value
            else:
                logger.warning(f"No value found for arrears band: {current_arrears_band}")
        
        # Insert data into the table
        for idx, header in enumerate(case_details_headers, start=1):
            value = data_mapping.get(header)
            if isinstance(value, (list, dict)):
                value = str(value)  # Convert lists or dicts to strings
            cell = ws.cell(row=start_row + idx, column=start_col + 1, value=value)
            cell.border = border_style['border']
            if header in ["Case ID", "Incident ID"]:
                cell.font = Font(bold=True)  # Make Case ID and Incident ID bold
        
        # Adjust column widths for the current table
        for col_idx in range(start_col, start_col + 2):  # Only adjust columns for the current table
            max_length = 0
            column_letter = chr(64 + col_idx)
            for row in range(start_row, start_row + len(case_details_headers) + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Case Details table created successfully.")
        return start_row + len(case_details_headers) + 1  # Return the next row after the table
    except Exception as e:
        logger.error(f"Failed to create Case Details table: {e}")
        sys.exit(1)

def create_contact_details_table(ws, case_data, start_row, start_col):
    """
    Create the Contact Details table in the worksheet.
    """
    try:
        logger.info("Creating Contact Details table...")
        contacts_headers = ["Mobile", "Email", "Home Phone", "Address"]

        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(contacts_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Contact Info")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
       # Write headers vertically
        for idx, header in enumerate(contacts_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        # Insert contact data vertically
        contacts = case_data.get("contact", [])
        for contact_idx, contact in enumerate(contacts, start=1):
            ws.cell(row=start_row + 1 + contact_idx, column=start_col, value=contact.get("mob")).border = border_style['border']
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 1, value=contact.get("email")).border = border_style['border']
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 2, value=contact.get("lan")).border = border_style['border']
            ws.cell(row=start_row + 1 + contact_idx, column=start_col + 3, value=contact.get("address")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(contacts_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(contacts) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Contact Details table created successfully.")
        return start_row + len(contacts) + 3
    except Exception as e:
        logger.error(f"Failed to create Contact Details table: {e}")
        sys.exit(1)

def create_remarks_table(ws, case_data, start_row, start_col):
    """
    Create the Remarks table in the worksheet.
    """
    try:
        logger.info("Creating Remarks table...")
        remark_headers = ["Remark", "Remark Added by", "Remark Added Date"]

        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(remark_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Remarks")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']        
        
       # Write headers vertically
        for idx, header in enumerate(remark_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        
        
        # Insert remarks data
        remarks = case_data.get("remark", [])
        for remark_idx, remark in enumerate(remarks, start=1):
            ws.cell(row=start_row + 1 + remark_idx, column=start_col, value=remark.get("remark")).border = border_style['border']
            ws.cell(row=start_row + 1 + remark_idx, column=start_col + 1, value=remark.get("remark_added_by")).border = border_style['border']
            ws.cell(row=start_row + 1 + remark_idx, column=start_col + 2, value=remark.get("remark_added_date")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(remark_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(remarks) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Remarks table created successfully.")
        return start_row + len(remarks) + 3
    except Exception as e:
        logger.error(f"Failed to create Remarks table: {e}")
        sys.exit(1)

def create_settlement_table(ws, settlements, start_row, start_col):
    """
    Create the Settlement table in the worksheet.
    """
    try:
        logger.info("Creating Settlement table...")
        settlement_headers = [
            "Settlement ID", "Case ID", "DRC Name", "RO Name", "Status", "Status reason",
            "Status DTM", "Settlement Type", "Settlement Amount", "Settlement Phase",
            "Settlement Created by", "Settlement Created DTM", "Last Monitoring DTM", "Remark"
        ]

        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(settlement_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Settlement Details")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(settlement_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        
        
        # Insert settlement data
        for settlement_idx, settlement in enumerate(settlements, start=1):
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col, value=settlement.get("settlement_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 1, value=settlement.get("case_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 2, value=settlement.get("drc_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 3, value=settlement.get("ro_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 4, value=settlement.get("settlement_status")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 5, value=settlement.get("status_reason")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 6, value=settlement.get("status_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 7, value=settlement.get("settlement_type")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 8, value=format_with_thousand_separator(settlement.get("settlement_amount"))).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 9, value=settlement.get("settlement_phase")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 10, value=settlement.get("created_by")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 11, value=settlement.get("created_on")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 12, value=settlement.get("last_monitoring_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + settlement_idx, column=start_col + 13, value=settlement.get("remark")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(settlement_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(settlements) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Settlement table created successfully.")
        return start_row + len(settlements) + 3
    except Exception as e:
        logger.error(f"Failed to create Settlement table: {e}")
        sys.exit(1)

def create_settlement_plan_table(ws, settlement_plans, start_row, start_col):
    """
    Create the Settlement Plan table in the worksheet.
    """
    try:
        logger.info("Creating Settlement Plan table...")
        settlement_plan_headers = [
            "Settlement ID", "Installment Sequence", "Installment Settle Amount",
            "Accumulated Amount", "Plan Date and Time"
        ]
        
        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(settlement_plan_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Settlement Plan")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(settlement_plan_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        
        
        # Insert settlement plan data
        for plan_idx, plan in enumerate(settlement_plans, start=1):
            ws.cell(row=start_row + 1 + plan_idx, column=start_col, value=plan.get("settlement_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 1, value=plan.get("installment_seq")).border = border_style['border']
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 2, value=format_with_thousand_separator(plan.get("installment_settle_amount"))).border = border_style['border']
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 3, value=format_with_thousand_separator(plan.get("accumulated_amount"))).border = border_style['border']
            ws.cell(row=start_row + 1 + plan_idx, column=start_col + 4, value=plan.get("plan_date")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(settlement_plan_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(settlement_plans) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Settlement Plan table created successfully.")
        return start_row + len(settlement_plans) + 3
    except Exception as e:
        logger.error(f"Failed to create Settlement Plan table: {e}")
        sys.exit(1)

def create_approve_table(ws, approve_data, start_row, start_col):
    """
    Create the Approve table in the worksheet.
    """
    try:
        logger.info("Creating Approve table...")
        approve_headers = ["Approved Process", "Approved By", "Approved On", "Remark"]
        
        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(approve_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Approve Details")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(approve_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        
        
        # Insert approve data
        for approve_idx, approve in enumerate(approve_data, start=1):
            ws.cell(row=start_row + 1 + approve_idx, column=start_col, value=approve.get("approved_process")).border = border_style['border']
            ws.cell(row=start_row + 1 + approve_idx, column=start_col + 1, value=approve.get("approved_by")).border = border_style['border']
            ws.cell(row=start_row + 1 + approve_idx, column=start_col + 2, value=approve.get("approved_on")).border = border_style['border']
            ws.cell(row=start_row + 1 + approve_idx, column=start_col + 3, value=approve.get("remark")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(approve_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(approve_data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Approve table created successfully.")
        return start_row + len(approve_data) + 3
    except Exception as e:
        logger.error(f"Failed to create Approve table: {e}")
        sys.exit(1)

def create_case_status_table(ws, case_status_data, start_row, start_col):
    """
    Create the Case Status table in the worksheet.
    """
    try:
        logger.info("Creating Case Status table...")
        case_status_headers = [
            "Case Status", "Status Reason", "Created DTM", "Created By",
            "Notified DTM", "Expire DTM"
        ]
       
        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(case_status_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Case Status")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(case_status_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
             
        # Insert case status data
        for status_idx, status in enumerate(case_status_data, start=1):
            ws.cell(row=start_row + 1 + status_idx, column=start_col, value=status.get("case_status")).border = border_style['border']
            ws.cell(row=start_row + 1 + status_idx, column=start_col + 1, value=status.get("status_reason")).border = border_style['border']
            ws.cell(row=start_row + 1 + status_idx, column=start_col + 2, value=status.get("created_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + status_idx, column=start_col + 3, value=status.get("created_by")).border = border_style['border']
            ws.cell(row=start_row + 1 + status_idx, column=start_col + 4, value=status.get("notified_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + status_idx, column=start_col + 5, value=status.get("expire_dtm")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(case_status_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(case_status_data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Case Status table created successfully.")
        return start_row + len(case_status_data) + 3
    except Exception as e:
        logger.error(f"Failed to create Case Status table: {e}")
        sys.exit(1)

def create_case_negotiation_table(ws, case_nego_data, start_row, start_col):
    """
    Create the Case Negotiation table in the worksheet.
    """
    try:
        logger.info("Creating Case Negotiation table...")
        case_negotiation_headers = [
            "Negotiation", "RO Request", "Negotiation Remarks"
        ]
        
        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(case_negotiation_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Case Negotiation")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(case_negotiation_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']

        
       
        
        # Insert case negotiation data
        for nego_idx, nego in enumerate(case_nego_data, start=1):
            ws.cell(row=start_row + 1 + nego_idx, column=start_col, value=nego.get("negotiation")).border = border_style['border']
            ws.cell(row=start_row + 1 + nego_idx, column=start_col + 1, value=nego.get("ro_request")).border = border_style['border']
            ws.cell(row=start_row + 1 + nego_idx, column=start_col + 2, value=nego.get("negotiation_remarks")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(case_negotiation_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(case_nego_data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Case Negotiation table created successfully.")
        return start_row + len(case_nego_data) + 3
    except Exception as e:
        logger.error(f"Failed to create Case Negotiation table: {e}")
        sys.exit(1)    

def create_ro_negotiation_table(ws, ro_negotiation_data, start_row, start_col):
    """
    Create the RO Negotiation table in the worksheet.
    """
    try:
        logger.info("Creating RO Negotiation table...")
        ro_negotiation_headers = [
            "DRC ID", "RO ID", "Created DTM", "Field Reason ID", "Field Reason", "Remark"
        ]
        
        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(ro_negotiation_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="RO Negotiation Details")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(ro_negotiation_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
             
        # Insert RO Negotiation data
        for nego_idx, nego in enumerate(ro_negotiation_data, start=1):
            ws.cell(row=start_row + 1 + nego_idx, column=start_col, value=nego.get("drc_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + nego_idx, column=start_col + 1, value=nego.get("ro_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + nego_idx, column=start_col + 2, value=nego.get("created_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + nego_idx, column=start_col + 3, value=nego.get("field_reason_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + nego_idx, column=start_col + 4, value=nego.get("field_reason")).border = border_style['border']
            ws.cell(row=start_row + 1 + nego_idx, column=start_col + 5, value=nego.get("remark")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(ro_negotiation_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(ro_negotiation_data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("RO Negotiation table created successfully.")
        return start_row + len(ro_negotiation_data) + 3
    except Exception as e:
        logger.error(f"Failed to create RO Negotiation table: {e}")
        sys.exit(1)

def create_drc_table(ws, drc_data, start_row, start_col):
    """
    Create the DRC table in the worksheet.
    """
    try:
        logger.info("Creating DRC table...")
        drc_headers = [
            "Order ID", "DRC ID", "DRC Name", "Created DTM", "DRC Status", "Status DTM",
            "Expire DTM", "Case Removal Remark", "Removed By", "Removed DTM",
            "DRC Selection Logic", "Case Distribution Batch ID"
        ]

        # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(drc_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="DRC Details")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(drc_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
        
        
        # Insert DRC data
        for drc_idx, drc in enumerate(drc_data, start=1):
            ws.cell(row=start_row + 1 + drc_idx, column=start_col, value=drc.get("order_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 1, value=drc.get("drc_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 2, value=drc.get("drc_name")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 3, value=drc.get("created_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 4, value=drc.get("drc_status")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 5, value=drc.get("status_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 6, value=drc.get("expire_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 7, value=drc.get("case_removal_remark")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 8, value=drc.get("removed_by")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 9, value=drc.get("removed_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 10, value=drc.get("drc_selection_logic")).border = border_style['border']
            ws.cell(row=start_row + 1 + drc_idx, column=start_col + 11, value=drc.get("case_distribution_batch_id")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(drc_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(drc_data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("DRC table created successfully.")
        return start_row + len(drc_data) + 3
    except Exception as e:
        logger.error(f"Failed to create DRC table: {e}")
        sys.exit(1)        

def create_recovery_officers_table(ws, recovery_officers_data, start_row, start_col):
    """
    Create the Recovery Officers table in the worksheet.
    """
    try:
        logger.info("Creating Recovery Officers table...")
        ro_headers = [
            "RO ID", "Assigned DTM", "Assigned By", "Removed DTM", "Case Removal Remark"
        ]
        
         # Use global styles from the STYLES dictionary
        main_header_style = STYLES['MainHeader_Style']
        sub_header_style = STYLES['SubHeader_Style']
        border_style = STYLES['Border_Style']

        # Create main header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(ro_headers) -1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value="Recovery Officers Details")
        main_header_cell.font = main_header_style['font']
        main_header_cell.fill = main_header_style['fill']
        main_header_cell.border = main_header_style['border']
        main_header_cell.alignment = main_header_style['alignment']
        
        # Write headers vertically
        for idx, header in enumerate(ro_headers, start=1):
            header_cell = ws.cell(row=start_row + 1, column=start_col + idx - 1, value=header)
            header_cell.font = sub_header_style['font']
            header_cell.fill = sub_header_style['fill']
            header_cell.border = sub_header_style['border']
            header_cell.alignment = sub_header_style['alignment']
        
       
        
        # Insert Recovery Officers data
        for ro_idx, ro in enumerate(recovery_officers_data, start=1):
            ws.cell(row=start_row + 1 + ro_idx, column=start_col, value=ro.get("ro_id")).border = border_style['border']
            ws.cell(row=start_row + 1 + ro_idx, column=start_col + 1, value=ro.get("assigned_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + ro_idx, column=start_col + 2, value=ro.get("assigned_by")).border = border_style['border']
            ws.cell(row=start_row + 1 + ro_idx, column=start_col + 3, value=ro.get("removed_dtm")).border = border_style['border']
            ws.cell(row=start_row + 1 + ro_idx, column=start_col + 4, value=ro.get("case_removal_remark")).border = border_style['border']
        
        # Adjust column widths
        for col in range(start_col, start_col + len(ro_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(recovery_officers_data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Recovery Officers table created successfully.")
        return start_row + len(recovery_officers_data) + 3
    except Exception as e:
        logger.error(f"Failed to create Recovery Officers table: {e}")
        sys.exit(1)

def create_case_details_sheet(wb, case_data, db):
    """
    Create the Case Details sheet with all tables.
    """
    try:
        logger.info("Creating Case Details sheet...")
        ws = wb.active
        ws.title = "Case Details"
        
        # Define starting row and column for the first table
        start_row, start_col = 2, 1 #change to current pointer
        
        # Create the Case Details table
        next_row = create_case_details_table(ws, case_data, start_row, start_col, db)
        
        # Add a two-row gap between the tables
        gap_row = next_row + 2
        
        # Create the Contact Details table
        next_row = create_contact_details_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Contact Info and Remarks tables
        gap_row = next_row + 1
        
        # Create the Remarks table
        next_row = create_remarks_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Remarks and Settlement tables
        gap_row = next_row + 1
        
        # Retrieve settlement data for the case
        case_id = case_data.get("case_id")
        settlements = get_settlement_data(db, case_id)
        
        # Create the Settlement table if settlement data exists
        if settlements:
            next_row = create_settlement_table(ws, settlements, gap_row, start_col)
            gap_row = next_row + 2  # Add a gap after the Settlement table
        
        # Retrieve settlement plan data for the case
        settlement_plans = get_settlement_plan_data(db, case_id)
        
        # Create the Settlement Plan table if settlement plan data exists
        if settlement_plans:
            next_row=create_settlement_plan_table(ws, settlement_plans, gap_row, start_col)
            gap_row = next_row + 2


        # Retrieve approve data from the case_data
        approve_data = case_data.get("approve", [])
        
        # Create the Approve table if approve data exists
        if approve_data:
            next_row = create_approve_table(ws, approve_data, gap_row, start_col) 
            gap_row = next_row + 2   
        
        # Retrieve case status data from the case_data
        case_status_data = case_data.get("case_status", [])
        
        # Create the Case Status table if case status data exists
        if case_status_data:
            next_row = create_case_status_table(ws, case_status_data, gap_row, start_col)
            gap_row = next_row + 2 

        # # Retrieve case negotiation data from the case_data
        # case_nego_data = case_data.get("case_nego", [])
        
        # # Create the Case Negotiation table if case negotiation data exists
        # if case_nego_data:
        #     create_case_negotiation_table(ws, case_nego_data, gap_row, start_col)

        # Retrieve RO Negotiation data from the case_data
        ro_negotiation_data = case_data.get("ro_negotiation", [])
        
        # Create the RO Negotiation table if RO Negotiation data exists
        if ro_negotiation_data:
            next_row = create_ro_negotiation_table(ws, ro_negotiation_data, gap_row, start_col)
            gap_row = next_row + 2

        # Retrieve DRC data from the case_data
        drc_data = case_data.get("drc", [])
        
        # Create the DRC table if DRC data exists
        if drc_data:
            next_row = create_drc_table(ws, drc_data, gap_row, start_col)    
            gap_row = next_row + 2

            # Retrieve Recovery Officers data from each DRC entry
            recovery_officers_data = []
            for drc in drc_data:
                recovery_officers_data.extend(drc.get("recovery_officers", []))
            
            # Create the Recovery Officers table if Recovery Officers data exists
            if recovery_officers_data:
                create_recovery_officers_table(ws, recovery_officers_data, gap_row, start_col)

        logger.info("Case Details sheet created successfully.")
        return ws
    except Exception as e:
        logger.error(f"Failed to create Case Details sheet: {e}")
        sys.exit(1)



def export_case_details(db, incident_id, output_path, collection_name):
    """
    Export case details to an Excel file.
    """
    try:
        logger.info(f"Exporting case details for Incident ID: {incident_id}")
        case_collection = db[collection_name]
        case_data = case_collection.find_one({"incident_id": incident_id})
        
        if not case_data:
            logger.error(f"No case details found for Incident ID: {incident_id}")
            sys.exit(1)
        
        logger.info(f"Case data found: {case_data}")

        # Get the case ID
        case_id = case_data.get("case_id")
        if not case_id:
            logger.error("Case ID not found in the case data.")
            sys.exit(1)

        # Generate a timestamp in the format YYYYMMDD_HHMMSS
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        # Create the output file name
        output_file_name = f"Case_{case_id}_{timestamp}.xlsx"
        output_file_save_location = os.path.join(output_path, output_file_name)

        # Check if the file already exists
        counter = 1
        while os.path.exists(output_file_save_location):
            output_file_name = f"Case_{case_id}_{timestamp}_{counter}.xlsx"
            output_file_save_location = os.path.join(output_path, output_file_name)
            counter += 1

        os.makedirs(os.path.dirname(output_file_save_location), exist_ok=True)
        
        excelworkbook = Workbook()
        ws = create_case_details_sheet(excelworkbook, case_data, db)
        
        # if not output_path.endswith('.xlsx'):
        #     output_path = os.path.join(output_path, 'case_details.xlsx')
        
        
        
        try:
            excelworkbook.save(output_file_save_location)
            logger.info(f"Case details exported to {output_file_save_location}")
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            sys.exit(1)
    except Exception as e:
        logger.error(f"Failed to export case details: {e}")
        sys.exit(1)

def main():
    """
    Main function to execute the case details export process.
    """
    try:
        logger.info("Starting case details export process...")
        config = get_config()
        db = connect_db(config)
        incident_id = 78910  # Replace with the actual incident ID you want to export
        export_path = config['EXCEL_EXPORT_PATHS']['WIN_DB']
        collection_name = config['COLLECTIONS']['CASE_DETAIL_COLLECTION_NAME']
        
        export_case_details(db, incident_id, export_path, collection_name)
        
        logger.info("Case details export process completed.")
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()